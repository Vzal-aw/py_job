from openpyxl import *

#二级文件中需要对比的数据
exe1_value = []
#一级文件中需要对比的数据
exe2_value = []
#二级文件中的id
exe3_value = []

#定义函数获取execl的值
def Getvalue(file,li_na,colu):
    wb = load_workbook(file)
    worksheet1 = wb.active
    for cell in worksheet1[colu]:
        li_na.append(cell.value)

file_name = input("请输入二级文件路径及名称：")
colu1 = input("请输入数据对比的列：")
file_name2 = input("请输入一级文件路径及名称：")
colu2 = input("请输入数据对比的列：")
#调用函数获取数据并传入对应列表中
Getvalue(file_name,exe1_value,colu1)
Getvalue(file_name2,exe2_value,colu2)
Getvalue(file_name,exe3_value,colu='A')


#获取一级文件中与二级文件中数据相同的列表下标
exe1_num = []
count = 0
for i in exe1_value:
    for j in exe2_value:
        if i == j:
            exe1_num.append(count)
    count +=1
b = set(exe1_num)
exe1_num = list(b)

#将二级文件与一级文件相同数据的二级文件id与数据写入新的execl中
wb = Workbook()
workbook2 = wb.active
num = 1
for lis_vlu in exe1_num:
    k = exe3_value[lis_vlu]
    g = exe1_value[lis_vlu]
    workbook2.cell(row=num,column=1).value=k
    workbook2.cell(row=num,column=2).value=g
    num += 1

wb.save('test1.xlsx')








