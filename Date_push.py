import openpyxl,sys,time
from openpyxl.styles import colors,Font,PatternFill
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

#路径
Chrom_path = r'C:\Users\Admin\AppData\Local\Google\Chrome\Application\chromedriver.exe'
url_path = r'http://25.212.136.36/isc_sso/login?service=http%3A%2F%2F25.212.136.36%3A80%2Fisc_sso%2Foauth2.0%2FcallbackAuthorize?oauth20_callbackUrl=http://25.212.136.36/?redirect=/identityManage/accountManage&response_type=token&scope=&client_id=10001&redirect_uri=http%3A%2F%2F25.212.136.36%2F%3Fredirect%3D%2FidentityManage%2FaccountManage&state=ad372e7d-bbca-4456-a32d-ec8e8f28be31'
Dirver = webdriver.Chrome(executable_path=Chrom_path)

#登录数字身份
def log_in():
    Dirver.implicitly_wait(15)
    Dirver.get(url_path)
    Dirver.maximize_window()
    Dirver.find_element_by_xpath('//input[@id="username"]').send_keys('tyqx')
    Dirver.find_element_by_xpath('//input[@id="password"]').send_keys('Tyqxpa#5167')
    Dirver.find_element_by_xpath('//input[@id="submit_login"]').click()
log_in()

#打开execl表
workbooks = openpyxl.load_workbook('C:\\Users\\Admin\\Desktop\\test.xlsx')
wb = workbooks.active
rowcount = 1
rows = wb.max_row
for data in wb.rows:
    for cell in data:
        username = wb.cell(row=rowcount,column=1)
    account = username.value
    print(account)

    #输入数据查询
    Dirver.implicitly_wait(20)
    time.sleep(3)
    add_in = Dirver.find_element_by_xpath('//div[@class="table-content el-col el-col-24"]/form/div[3]/div/div/input[@type="text"]')
    add_in.send_keys(account)
    time.sleep(2)
    #加入限制条件，解决子集查询重复点击问题
    if rowcount == 1:
        Dirver.find_element_by_xpath('//span[@class="el-checkbox__inner"]').click()
    rowcount += 1
    Dirver.find_element_by_xpath('//div[@class="table-content el-col el-col-24"]/form/div[5]/div/button[@type="button"]').click()
    add_in.clear()
    time.sleep(8)

    #筛选数据，
    text_li = []
    Sument = Dirver.find_elements_by_class_name('el-table__row')
    for i in Sument:
        text = i.text
        text1 = text.split()
        if account not in text1:
            continue
        elif account in text1:
            text_li.append(text1[:2])
    serial_num = text_li[0]

    #账号禁用启用
    def data_push():
        Dirver.implicitly_wait(15)
        Dirver.find_element_by_xpath('//tbody/tr[{}]/td[2]/div[1]/label[@role="checkbox"]'.format(serial_num[0])).click()
        Dirver.find_element_by_xpath('//div[@class="table-content el-col el-col-24"]/form/div[last()]/div/div/div/input[@type="text"]').click()
        time.sleep(2)
        Dirver.find_element_by_xpath('//div[@x-placement="bottom-start"]/div/div/ul/li[7]').click()
        Dirver.find_element_by_xpath('//div[@aria-label="账号禁用"]/div[3]/div/button[@type="button"]').click()
        time.sleep(3)
        Dirver.find_element_by_xpath('//tbody/tr[{}]/td[2]/div[1]/label[@role="checkbox"]'.format(serial_num[0])).click()
        Dirver.find_element_by_xpath('//div[@class="table-content el-col el-col-24"]/form/div[last()]/div/div/div/input[@type="text"]').click()
        time.sleep(3)
        Dirver.find_element_by_xpath('//div[@x-placement="bottom-start"]/div/div/ul/li[8]').click()
        Dirver.find_element_by_xpath('//div[@aria-label="账号启用"]/div[3]/div/button[@type="button"]').click()
        time.sleep(2)
    data_push()










