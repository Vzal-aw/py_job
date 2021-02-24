import openpyxl,sys,time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

"""启动谷歌浏览器并打开对应网站"""
url = r'http://iscsso.sgcc.com.cn/isc_sso/login?service=http%3A%2F%2Fiscmp.sgcc.com.cn%2Fisc_mp%2Fframework%2Fdesktop%2Findex.jsp#'
driver_path = r'C:\Users\Admin\AppData\Local\Google\Chrome\Application\chromedriver'
driver = webdriver.Chrome(executable_path=driver_path)
driver.get(url)
driver.maximize_window()


"""用户登录"""
Company = driver.find_element_by_xpath("//input[@id='department']")
Company_1 = driver.find_element_by_class_name("tab_item2")
Company_2 = driver.find_element_by_xpath("//a[@data='hn']")
ActionChains(driver).click(Company).click(Company_1).click(Company_2).perform()
Username = driver.find_element_by_xpath("//li//input[@id='username']")
Username.send_keys('ml')
Passwor = driver.find_element_by_xpath("//li//input[@id='password']")
Passwor.send_keys('hn@4306!')
driver.find_element_by_xpath("//input[@name='submit_login']").send_keys(Keys.ENTER)

"""自动解锁"""
driver.find_element_by_xpath("//span[@id='treeDemo_1_switch']").click()
driver.find_element_by_xpath("//span[@id='treeDemo_2_ico']").click()
time.sleep(8)
test = driver.find_element_by_xpath("//*[@id='tabs']/div[2]/div[2]/div/iframe")
driver.switch_to.frame(test)

"""读取execl表数据解锁"""
workbooks = openpyxl.load_workbook('C:\\Users\\Admin\\Desktop\\test.xlsx')
wb = workbooks.active
rowcount = 1
rows = wb.max_row
for i in wb.rows:
    for cell in i:
        firstname = wb.cell(row=rowcount,column=1)
    first_value = firstname.value
    if rowcount > rows:
        break
    rowcount += 1
    driver.find_element_by_xpath('//tbody[1]/tr[1]/td[2]/div/input').send_keys(first_value)
    driver.find_element_by_xpath('//span[@id="img"]').click()
    driver.find_element_by_xpath('//input[@type="button" and @value="查询" ]').click()
    time.sleep(3)
    Stute =  driver.find_element_by_xpath('//table[@cellspacing="0"]/tbody[1]/tr[1]/td[@align="left" and @id="state"]').text
    print(Stute)
    if Stute == "启用":
        driver.find_element_by_xpath('//tbody[1]/tr[1]/td[2]/div/input').clear()
        continue
    if Stute == "锁定":
        driver.find_element_by_xpath('//table[@cellspacing="0"]/tbody[1]/tr[1]/td[2]/div/span[@id="checkBox"]').click()
        driver.find_element_by_xpath('//ul[@class="toolBar"]/li[last()-2]/a[1]/span[text()="登录解锁"]').click()
        driver.find_element_by_xpath('//tbody[1]/tr[1]/td[2]/div/input').clear()
        time.sleep(3)

    # if Stute == "禁用":
    #     driver.find_element_by_xpath('//table[@cellspacing="0"]/tbody[1]/tr[1]/td[2]/div/span[@id="checkBox"]').click()
    #     driver.find_element_by_xpath('//ul[@class="toolBar"]/li[last()-6]/a/span').click()
    #     test = driver.find_element_by_xpath('//dir[@class="container mx resizable window active"]')
    #     print(test)
    #     driver.find_element_by_xpath('//tbody[1]/tr[1]/td[2]/div/input').clear()
























