import openpyxl,sys,time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

iedirver_path = r'C:\Program Files\Internet Explorer\IEDriverServer.exe'
url = r'http://uums-a.hn.sgcc.com.cn/uums/login.view?method=getMenu&currentIndex=50'
Dirver = webdriver.Ie(executable_path=iedirver_path)


#运行时间增加装饰器
def time_out(f):
    def add_time(*args,**kwargs):
        time.sleep(3)
        ret = f(*args,**kwargs)
        return ret
    return add_time()


#统一用户登录
Dirver.get(url)
Submit = Dirver.find_element_by_xpath('//input[@name="commit"]')
Dirver.find_element_by_xpath('//input[@name="username"]').send_keys('admin')
Dirver.find_element_by_xpath('//input[@name="password"]').send_keys('n0v3ll??')
loginto = Dirver.find_element_by_xpath('//input[@name="commit"]').click()


#用户维护
@time_out
def add_cread():
    Dirver.find_element_by_xpath('//ul/li[@id="two2"]').click()
    time.sleep(1)
    Dirver.find_element_by_xpath('//div[@id="leftmenu_0"]/li[@id="twomenu1"]/a[@class="topmenu"]').click()


"""通过读取execl 表中的数据判断对应的组织机构"""

workbooks = openpyxl.load_workbook('C:\\Users\\Admin\\Desktop\\account_cbuild.xlsx')
wb = workbooks.active
rowcount = 1
rows = wb.max_row
for i in wb.rows:
    for cell in i:
        firstname = wb.cell(row=rowcount,column=1)
        account = wb.cell(row=rowcount,column=2)
        org = wb.cell(row=rowcount,column=3)

    first_value = firstname.value
    last_value = account.value
    org_value = org.value
    """获取exel表中组织机构并添加至列表中"""
    org = phon_value.split("|")
    print(org)
    """读取列表中的组织机构名称判断组织机构位置"""
    if len(org) >= 1:
        Dirver.find_element_by_link_text(org[0]).click()
    if len(org) >= 2:
        Dirver.find_element_by_xpath('')
    if len(org) >= 3:
        Dirver.find_element_by_xpath('')

    if rowcount > rows:
        break

    rowcount += 1


















