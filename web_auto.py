import openpyxl,sys,time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

"""启动谷歌浏览器并打开对应网站"""
url = r'http://iscsso.sgcc.com.cn/isc_sso/login?service=http%3A%2F%2Fiscmp.sgcc.com.cn%2Fisc_mp%2Fframework%2Fdesktop%2Findex.jsp#'
driver_path = r'C:Users\Admin\AppData\Local\Google\Chrome\Application\chromedriver.exe'
driver = webdriver.Chrome(executable_path=driver_path)
driver.get(url)

"""ActionChains方法选取组织机构"""
Company = driver.find_element_by_xpath("//input[@id='department']")
Company_1 = driver.find_element_by_class_name("tab_item2")
Company_2 = driver.find_element_by_xpath("//a[@data='hn']")
ActionChains(driver).click(Company).click(Company_1).click(Company_2).perform()

"""输入用户名及密码进行登录"""
Username = driver.find_element_by_xpath("//li//input[@id='username']")
Username.send_keys('ml')
Passwor = driver.find_element_by_xpath("//li//input[@id='password']")
Passwor.send_keys('hn@4306!')
driver.find_element_by_xpath("//input[@name='Submit']").send_keys(Keys.ENTER)

"""读取execl表数据并传入"""
mb = openpyxl.load_workbook('C:Users\\Admin\\Desktop\\test.xlsx')
wb = mb.get_sheet_by_name('Sheet1')


"""自动完成用户创建"""
plus = driver.find_element_by_xpath("//span[@id='treeDemo_1_switch']")
ActionChains(driver).click(plus).perform()
time.sleep(1)
user_admin = driver.find_element_by_xpath("//span[@id='treeDemo_2_ico']")
ActionChains(driver).click(user_admin).perform()
time.sleep(5)
test = driver.find_element_by_xpath("//*[@id='tabs']/div[2]/div[2]/div/iframe")
driver.switch_to.frame(test)


"""读取execl表中的数据"""
workbooks = openpyxl.load_workbook('C:\\Users\\Admin\\Desktop\\account_cbuild.xlsx')
wb = workbooks.active
rowcount = 1
rows = wb.max_row
for i in wb.rows:
    for cell in i:
        firstname = wb.cell(row=rowcount,column=1)
        account = wb.cell(row=rowcount,column=2)
        phon = wb.cell(row=rowcount,column=3)


    first_value = firstname.value
    last_value = account.value
    phon_value = phon.value
    if rowcount > rows:
        break
    rowcount += 1
    time.sleep(2)
    user_create = driver.find_element_by_xpath("//span[text()='注册正式用户']")
    ActionChains(driver).click(user_create).perform()
    add_user = driver.find_element_by_xpath("//span[text()='添加']")
    ActionChains(driver).click(add_user).perform()
    """查询用户"""
    input_name = driver.find_element_by_xpath("//input[@type='text']").send_keys(first_value)
    index_click = driver.find_element_by_xpath("//span[@id='img']")

    ActionChains(driver).click(index_click).perform()
    time.sleep(3)
    win = driver.find_element_by_xpath("//td[@id='checkBoxTD']/div[1]")
    ActionChains(driver).click(win).perform()
    confirm = driver.find_element_by_xpath("//span[text()='确认所选用户']")
    ActionChains(driver).click(confirm).perform()
    time.sleep(1)
    """输入账号"""
    name_account = driver.find_element_by_xpath("//td[@id='namecode'][not(@style)]")
    ActionChains(driver).click(name_account).perform()
    username = driver.find_element_by_xpath("//td[@id='namecode'][not(@style)]/div/input[1]")
    time.sleep(0.5)
    username.send_keys(last_value)
    """输入邮箱地址"""
    email_account = driver.find_element_by_xpath("//td[@id='email'][not(@style)]")
    ActionChains(driver).click(email_account).perform()
    email = driver.find_element_by_xpath("//td[@id='email'][not(@style)]/div/input[1]")
    time.sleep(0.5)
    email.send_keys("{}@hn.sgcc.com.cn".format(last_value))
    """输入手机号"""
    phon_account = driver.find_element_by_xpath("//td[@id='tel'][not(@style)]")
    driver.execute_script("arguments[0].scrollIntoView();", phon_account)
    ActionChains(driver).click(phon_account).perform()
    phon_num = driver.find_element_by_xpath("//td[@id='tel'][not(@style)]/div/input[1]")
    time.sleep(0.5)
    phon_num.send_keys(phon_value)
    save = driver.find_element_by_xpath("//li[@id='save']/a/span[1]")
    ActionChains(driver).click(save).perform()





